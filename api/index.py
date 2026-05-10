"""
DCF Valuation Tool — Flask backend
Serves the React frontend and handles:
  - /api/extract  → forwards uploaded docs to Anthropic API, returns JSON
  - /api/generate → writes user inputs into Valuation_Template.xlsx, returns the file
"""

import os
import io
import json
import base64
import tempfile
import copy
from datetime import datetime

import requests
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook

app = Flask(__name__, static_folder="static", template_folder="templates")
CORS(app)

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "Valuation_Template.xlsx")

# ── Serve the React SPA ──────────────────────────────────────────────────────
@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_spa(path):
    if path and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.template_folder, "index.html")


# ── /api/extract  ─────────────────────────────────────────────────────────────
@app.route("/api/extract", methods=["POST"])
def extract():
    """Receive uploaded files + context, call Claude, return extracted JSON."""
    description = request.form.get("description", "deep-tech medtech startup")
    files = request.files.getlist("files")

    content_blocks = []

    for f in files:
        raw = f.read()
        mime = f.mimetype or "application/octet-stream"
        filename = f.filename or "upload"

        if mime == "application/pdf":
            b64 = base64.b64encode(raw).decode()
            content_blocks.append({
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64}
            })
        elif mime.startswith("image/"):
            b64 = base64.b64encode(raw).decode()
            content_blocks.append({
                "type": "image",
                "source": {"type": "base64", "media_type": mime, "data": b64}
            })
        else:
            # Try to decode as text (xlsx handled client-side for extraction)
            try:
                text = raw.decode("utf-8", errors="replace")
            except Exception:
                text = f"[Binary file: {filename}]"
            content_blocks.append({
                "type": "text",
                "text": f"File: {filename}\n{text[:12000]}"
            })

    extraction_prompt = f"""
You are a senior financial analyst. Extract ALL financial data from the documents above.
Company context: {description}

Return ONLY valid JSON — no markdown fences, no explanation. Use null for values not found.
Monetary values as numbers (no currency symbols). Year arrays = 11 values for 2025–2035.

{{
  "currency": "CHF",
  "company_stage": null,
  "funding_raised_total": null,
  "revenue_year1": null,
  "revenue_year2": null,
  "revenue_year3": null,
  "revenue_year5": null,
  "revenue_growth_rate": null,
  "gross_margin": null,
  "ebitda_margin": null,
  "burn_rate_monthly": null,
  "runway_months": null,
  "tam_size": null,
  "deal1_signing_year": null,
  "deal1_upfront_fee": null,
  "deal1_royalty_rate": null,
  "deal1_partner_revenue": [null,null,null,null,null,null,null,null,null,null,null],
  "deal2_codev_fee": [null,null,null,null,null,null,null,null,null,null,null],
  "deal2_royalty_rate": null,
  "deal2_partner_revenue": [null,null,null,null,null,null,null,null,null,null,null],
  "ftes": [null,null,null,null,null,null,null,null,null,null,null],
  "avg_fte_cost": null,
  "capex": [null,null,null,null,null,null,null,null,null,null,null],
  "wacc_rf": null,
  "wacc_erp": null,
  "tax_rate": null,
  "terminal_growth": null,
  "milestone_fda_year": null,
  "notes": "List every specific number found and its source section."
}}"""

    content_blocks.append({"type": "text", "text": extraction_prompt})

    if not ANTHROPIC_API_KEY:
        return jsonify({"error": "ANTHROPIC_API_KEY not set on server"}), 500

    resp = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "Content-Type": "application/json",
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
        json={
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 2000,
            "messages": [{"role": "user", "content": content_blocks}],
        },
        timeout=60,
    )

    data = resp.json()
    raw_text = (data.get("content") or [{}])[0].get("text", "")
    raw_text = raw_text.replace("```json", "").replace("```", "").strip()
    try:
        extracted = json.loads(raw_text)
    except Exception:
        extracted = {"notes": "Could not parse extraction output."}

    return jsonify(extracted)


# ── /api/generate  ────────────────────────────────────────────────────────────
@app.route("/api/generate", methods=["POST"])
def generate():
    """Receive form values + extracted data, fill template, return xlsx."""
    body = request.get_json(force=True)
    form = body.get("form", {})
    ex = body.get("extracted", {})
    wacc = body.get("wacc", {})
    comps = body.get("comps", [])
    bench = body.get("bench", {})
    sector_name = body.get("sectorName", "MedTech")

    wb = load_workbook(TEMPLATE_PATH)
    today = datetime.now().strftime("%B %Y")

    def inj(ws, ref, val):
        """Inject value only into non-formula cells."""
        cell = ws[ref]
        if cell.data_type == "f" or (hasattr(cell, "value") and isinstance(cell.value, str) and cell.value.startswith("=")):
            return
        cell.value = val

    def inj_row(ws, row, val):
        for col in "CDEFGHIJKLM":
            cell = ws[f"{col}{row}"]
            if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = val

    def inj_arr(ws, row, arr11):
        if not arr11:
            return
        for i, col in enumerate("CDEFGHIJKLM"):
            if arr11[i] is not None:
                cell = ws[f"{col}{row}"]
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.value = arr11[i]

    # ── Cover ──────────────────────────────────────────────────────────────
    cvr = wb["Cover"]
    inj(cvr, "C2", form.get("company", ""))
    inj(cvr, "C3", form.get("description", "")[:150])
    inj(cvr, "C4", form.get("revenueModel", ""))
    inj(cvr, "C5", f"v1.0 — Generated {today}")
    inj(cvr, "C6", "2025 – 2035")
    inj(cvr, "C7", today)
    inj(cvr, "C8", "Generated by DCF Valuation Tool")

    # ── Assumptions ────────────────────────────────────────────────────────
    ass = wb["Assumptions"]
    inj(ass, "B1", f"{form.get('company', '')} — DCF MODEL | ASSUMPTIONS")

    if wacc:
        inj_row(ass, 4,  wacc.get("rf"))
        inj_row(ass, 5,  wacc.get("erp"))
        inj_row(ass, 6,  wacc.get("betaU"))
        inj_row(ass, 8,  wacc.get("kd"))
        inj_row(ass, 9,  wacc.get("tax"))
        inj_row(ass, 10, wacc.get("de"))
        inj_row(ass, 49, wacc.get("g"))

    # Revenue assumptions
    if ex.get("deal1_upfront_fee"):
        inj(ass, "C16", ex["deal1_upfront_fee"])
    if ex.get("deal1_royalty_rate"):
        inj_row(ass, 17, ex["deal1_royalty_rate"])
    inj_arr(ass, 18, ex.get("deal1_partner_revenue"))
    inj_arr(ass, 23, ex.get("deal2_codev_fee"))
    if ex.get("deal2_royalty_rate"):
        inj_row(ass, 24, ex["deal2_royalty_rate"])
    inj_arr(ass, 25, ex.get("deal2_partner_revenue"))

    # OPEX assumptions
    inj_arr(ass, 35, ex.get("ftes"))
    if ex.get("avg_fte_cost"):
        inj_row(ass, 36, ex["avg_fte_cost"])
    inj_arr(ass, 44, ex.get("capex"))

    # Terminal value
    if ex.get("terminal_growth"):
        inj_row(ass, 49, ex["terminal_growth"])

    # ── Scenarios ──────────────────────────────────────────────────────────
    scen = wb["Scenarios"]
    inj(scen, "B1", f"SCENARIO ANALYSIS — BEAR / BASE / BULL | {form.get('company', '')}")

    base_rev1 = ex.get("revenue_year1") or 500000
    base_rev3 = ex.get("revenue_year3") or base_rev1 * 3
    wacc_val = wacc.get("wacc", 0.18)
    g_val = wacc.get("g", 0.015)

    scen["C4"].value = "3.0%"
    scen["D4"].value = "5.0%"
    scen["E4"].value = "8.0%"
    scen["C5"].value = round(base_rev1 * 0.5)
    scen["D5"].value = round(base_rev1)
    scen["E5"].value = round(base_rev1 * 2.5)
    scen["C6"].value = 1
    scen["D6"].value = 2
    scen["E6"].value = 4
    scen["C7"].value = f"{(wacc_val + 0.03) * 100:.1f}%"
    scen["D7"].value = f"{wacc_val * 100:.1f}%"
    scen["E7"].value = f"{(wacc_val - 0.02) * 100:.1f}%"
    scen["C8"].value = "1.0%"
    scen["D8"].value = f"{g_val * 100:.1f}%"
    scen["E8"].value = "2.5%"
    scen["C9"].value = "8.0x"
    scen["D9"].value = "12.0x"
    scen["E9"].value = "18.0x"
    scen["C10"].value = round(base_rev3 * 0.4)
    scen["D10"].value = round(base_rev3)
    scen["E10"].value = round(base_rev3 * 2.2)

    # ── Add Comparables sheet ──────────────────────────────────────────────
    if "Comparables" in wb.sheetnames:
        del wb["Comparables"]
    ws_comp = wb.create_sheet("Comparables", 0)
    ws_comp["B1"] = f"COMPARABLE COMPANIES — {sector_name.upper()} | {form.get('company', '')} DCF"
    ws_comp["B3"] = "Company"
    ws_comp["C3"] = "Ticker"
    ws_comp["D3"] = "Country"
    ws_comp["E3"] = "Gross Margin %"
    ws_comp["F3"] = "EBITDA Margin %"
    ws_comp["G3"] = "Beta (unlev.)"
    ws_comp["H3"] = "Rev. Growth %"

    if bench:
        ws_comp["B5"] = "SECTOR MEDIANS"
        ws_comp["E5"] = f"{bench.get('gm', 0):.1f}%"
        ws_comp["F5"] = f"{bench.get('eb', 0):.1f}%"
        ws_comp["G5"] = f"{bench.get('b', 0):.3f}"
        ws_comp["H5"] = f"{bench.get('rg', 0):.1f}%"

    for i, c in enumerate(comps):
        row = 7 + i
        ws_comp[f"B{row}"] = c.get("n", "")
        ws_comp[f"C{row}"] = c.get("t", "")
        ws_comp[f"D{row}"] = c.get("c", "")
        ws_comp[f"E{row}"] = f"{c.get('gm', 0):.1f}%" if c.get("gm") else "—"
        ws_comp[f"F{row}"] = f"{c.get('eb', 0):.1f}%" if c.get("eb") else "—"
        ws_comp[f"G{row}"] = f"{c.get('b', 0):.3f}" if c.get("b") else "—"
        ws_comp[f"H{row}"] = f"{c.get('rg', 0):.1f}%" if c.get("rg") else "—"

    # ── Save to buffer & return ────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = form.get("company", "Startup").replace(" ", "_").replace("/", "-")
    filename = f"Valuation_{safe_name}.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
