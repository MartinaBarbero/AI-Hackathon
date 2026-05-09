"""
YFINANCE TICKER VERIFICATION — Medtech Comparable Database
Run with: python test_yfinance.py
Install:  pip install yfinance openpyxl

Tests all 66 active tickers across 10 sub-sectors.
Outputs a color-coded Excel file with results.
Delisted tickers (marked †) are expected to fail — noted in output.
"""

import yfinance as yf
import time
from datetime import datetime

TICKERS = {
    # SUB-SECTOR 1 — Neurovascular & Neurotech
    "Penumbra (US)":               ("PEN",        "green",  False),
    "Nyxoah (BE)":                 ("NYXH",       "yellow", False),
    "Silk Road Medical (US)†":     ("SILK",       "yellow", True),
    "ClearPoint Neuro (US)":       ("CLPT",       "yellow", False),
    "InspireMD (US)":              ("NSPR",       "yellow", False),
    # SUB-SECTOR 2 — Orthopedics & Musculoskeletal
    "BONESUPPORT (SE)":            ("BONEX.ST",   "green",  False),
    "SI-BONE (US)":                ("SIBN",       "green",  False),
    "NuVasive (US)†":              ("NUVA",       "green",  True),
    "Bioretec (FI)":               ("BRETEC.HE",  "yellow", False),
    "Alphatec Holdings (US)":      ("ATEC",       "yellow", False),
    "Orthofix Medical (US)":       ("OFIX",       "yellow", False),
    # SUB-SECTOR 3 — Wearable & Drug Delivery
    "Convatec Infusion Care (UK)": ("CTEC.L",     "yellow", False),
    "Tandem Diabetes Care (US)":   ("TNDM",       "yellow", False),
    "Embecta Corp (US)":           ("EMBC",       "yellow", False),
    # SUB-SECTOR 4 — Cardiovascular & Interventional
    "XVIVO Perfusion (SE)":        ("XVIVO.ST",   "green",  False),
    "AtriCure (US)":               ("ATRC",       "green",  False),
    "LeMaitre Vascular (US)":      ("LMAT",       "green",  False),
    "Artivion (US)":               ("AORT",       "green",  False),
    "AngioDynamics (US)":          ("ANGO",       "yellow", False),
    # SUB-SECTOR 5 — Diagnostics & IVD
    "Stratec SE (DE)":             ("SBS.DE",     "green",  False),
    "EKF Diagnostics (UK)":        ("EKF.L",      "green",  False),
    "Boule Diagnostics (SE)":      ("BOUL.ST",    "yellow", False),
    "Tecan Group (CH)":            ("TECN.SW",    "yellow", False),
    "Vitrolife (SE)":              ("VITR.ST",    "yellow", False),
    "Tristel (UK)":                ("TSTL.L",     "yellow", False),
    "genedrive (UK)":              ("GDR.L",      "yellow", False),
    "Photocure (NO)":              ("PHO.OL",     "yellow", False),
    "Veracyte (US)":               ("VCYT",       "yellow", False),
    "CareDx (US)":                 ("CDNA",       "yellow", False),
    "Castle Biosciences (US)":     ("CSTL",       "yellow", False),
    # SUB-SECTOR 6 — Surgical Robotics & MIS
    "Ambu (DK)":                   ("AMBU-B.CO",  "green",  False),
    "Creo Medical (UK)":           ("CREO.L",     "yellow", False),
    "PROCEPT BioRobotics (US)":    ("PRCT",       "yellow", False),
    "Mazor Robotics (IL)†":        ("MZOR",       "yellow", True),
    "MAKO Surgical (US)†":         ("MAKO",       "yellow", True),
    "Stereotaxis (US)":            ("STXS",       "yellow", False),
    # SUB-SECTOR 7 — Digital Health & SaMD
    "RaySearch Labs (SE)":         ("RAY-B.ST",   "green",  False),
    "Sectra (SE)":                 ("SECT-B.ST",  "yellow", False),
    "Ascom Holding (CH)":          ("ASCN.SW",    "yellow", False),
    "Median Technologies (FR)":    ("ALMDT.PA",   "yellow", False),
    "Polarean Imaging (UK)":       ("PLLWF",      "yellow", False),
    "iRhythm Technologies (US)":   ("IRTC",       "yellow", False),
    "Butterfly Network (US)":      ("BFLY",       "yellow", False),
    "Nano-X Imaging (IL)":         ("NNOX",       "yellow", False),
    "Hyperfine (US)":              ("HYPR",       "yellow", False),
    # SUB-SECTOR 8 — Ophthalmology & Photonics
    "Revenio Group (FI)":          ("REG1V.HE",   "green",  False),
    "STAAR Surgical (US)":         ("STAA",       "green",  False),
    "Optomed (FI)":                ("OPTOMED.HE", "yellow", False),
    "El.En. SpA (IT)":             ("ELN.MI",     "yellow", False),
    "IBA (BE)":                    ("IBAB.BR",    "yellow", False),
    "Lumibird (FR)":               ("LBIRD.PA",   "yellow", False),
    "Glaukos (US)":                ("GKOS",       "yellow", False),
    "LENSAR (US)":                 ("LNSR",       "yellow", False),
    # SUB-SECTOR 9 — Regenerative Medicine & Wound Care
    "Vericel (US)":                ("VCEL",       "green",  False),
    "Eckert & Ziegler (DE)":       ("EUZ.DE",     "yellow", False),
    "Convatec Wound Care (UK)":    ("CTEC.L",     "yellow", False),
    "Axogen (US)":                 ("AXGN",       "yellow", False),
    "Organogenesis (US)":          ("ORGO",       "yellow", False),
    # SUB-SECTOR 10 — Monitoring & Implantables
    "LivaNova (UK/Nasdaq)":        ("LIVN",       "green",  False),
    "Dragerwerk (DE)":             ("DRW3.DE",    "yellow", False),
    "Bactiguard (SE)":             ("BACTI-B.ST", "yellow", False),
    "Inspiration Healthcare (UK)": ("IHC.L",      "yellow", False),
    "Elekta (SE)":                 ("EKTA-B.ST",  "yellow", False),
    "Inspire Medical (US)":        ("INSP",       "yellow", False),
    "Senseonics (US)":             ("SENS",       "yellow", False),
    "CVRx (US)":                   ("CVRX",       "yellow", False),
}

INFO_METRICS = {
    "Gross Margin":     "grossMargins",
    "EBITDA Margin":    "ebitdaMargins",
    "Op. Margin":       "operatingMargins",
    "Revenue Growth":   "revenueGrowth",
    "Beta":             "beta",
    "Debt/Equity":      "debtToEquity",
    "Tax Rate":         "effectiveTaxRate",
    "Market Cap":       "marketCap",
    "Country":          "country",
    "Currency":         "currency",
}

CRITICAL = {"beta", "grossMargins", "country", "currency"}

GREEN = "\033[92m"; YELLOW = "\033[93m"; RED = "\033[91m"
CYAN = "\033[96m";  BOLD = "\033[1m";    RESET = "\033[0m"


def pct(v):
    return f"{v*100:.1f}%" if v is not None and abs(v) < 10 else (f"{v:.1f}%" if v else None)

def fmt_cap(v):
    return f"${v/1e9:.2f}B" if v else None


def test_ticker(name, ticker, verdict, is_delisted):
    r = {"name": name, "ticker": ticker, "verdict": verdict, "delisted": is_delisted,
         "status": None, "metrics": {}, "financials_ok": False, "cashflow_ok": False, "issues": []}

    if is_delisted:
        r["status"] = "delisted"
        r["issues"].append("Expected failure. Use SEC EDGAR for historical data.")
        return r

    try:
        t = yf.Ticker(ticker)
        info = t.info
        missing_critical = []

        for label, key in INFO_METRICS.items():
            raw = info.get(key)
            if raw in (None, "", "N/A") and key in CRITICAL:
                missing_critical.append(label)
            if key in ("grossMargins","ebitdaMargins","operatingMargins","revenueGrowth","effectiveTaxRate"):
                r["metrics"][label] = pct(raw)
            elif key == "marketCap":
                r["metrics"][label] = fmt_cap(raw)
            else:
                r["metrics"][label] = raw

        try:
            fin = t.financials
            r["financials_ok"] = fin is not None and not fin.empty
            if r["financials_ok"]:
                r["metrics"]["Financials (years)"] = fin.shape[1]
        except Exception as e:
            r["issues"].append(f"financials: {e}")

        try:
            cf = t.cashflow
            r["cashflow_ok"] = cf is not None and not cf.empty
        except Exception as e:
            r["issues"].append(f"cashflow: {e}")

        if missing_critical:
            r["status"] = "partial"
            r["issues"].append(f"Critical missing: {', '.join(missing_critical)}")
        elif not r["financials_ok"]:
            r["status"] = "partial"
            r["issues"].append("Financials empty — use annual report")
        else:
            r["status"] = "ok"

    except Exception as e:
        r["status"] = "fail"
        r["issues"].append(str(e))

    return r


def print_result(r):
    icon = "🟢" if r["verdict"] == "green" else "🟡"
    if r["status"] == "ok":       s = f"{GREEN}✅ OK{RESET}"
    elif r["status"] == "partial": s = f"{YELLOW}⚠️  PARTIAL{RESET}"
    elif r["status"] == "delisted":s = f"{CYAN}†  DELISTED{RESET}"
    else:                          s = f"{RED}❌ FAIL{RESET}"
    print(f"  {icon} {BOLD}{r['name']}{RESET} [{r['ticker']}] → {s}")
    if r["status"] in ("ok","partial"):
        for k, v in r["metrics"].items():
            if v not in (None, False, "None", ""):
                print(f"     {k:<25} {v}")
    for issue in r["issues"]:
        print(f"     {YELLOW}→ {issue}{RESET}")


def save_excel(results):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("\n[skip] pip install openpyxl to enable Excel export")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    FILLS = {"ok": PatternFill("solid",fgColor="C6EFCE"),
             "partial": PatternFill("solid",fgColor="FFEB9C"),
             "delisted": PatternFill("solid",fgColor="BDD7EE"),
             "fail": PatternFill("solid",fgColor="FFC7CE")}
    H_FILL = PatternFill("solid",fgColor="1F4E79")
    H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    N_FONT = Font(name="Arial", size=9)
    headers = ["Verdict","Name","Ticker","Status","Gross Margin","EBITDA Margin",
               "Op. Margin","Rev. Growth","Beta","D/E","Tax Rate","Market Cap",
               "Financials OK","Cashflow OK","Country","Currency","Issues"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    STATUS_LABEL = {"ok":"✅ OK","partial":"⚠️ PARTIAL","delisted":"† DELISTED","fail":"❌ FAIL"}
    for row, r in enumerate(results, 2):
        vals = ["🟢" if r["verdict"]=="green" else "🟡", r["name"], r["ticker"],
                STATUS_LABEL.get(r["status"],"?"),
                r["metrics"].get("Gross Margin","—"), r["metrics"].get("EBITDA Margin","—"),
                r["metrics"].get("Op. Margin","—"), r["metrics"].get("Revenue Growth","—"),
                r["metrics"].get("Beta","—"), r["metrics"].get("Debt/Equity","—"),
                r["metrics"].get("Tax Rate","—"), r["metrics"].get("Market Cap","—"),
                "✅" if r["financials_ok"] else "❌", "✅" if r["cashflow_ok"] else "❌",
                r["metrics"].get("Country","—"), r["metrics"].get("Currency","—"),
                " | ".join(r["issues"])]
        fill = FILLS.get(r["status"], FILLS["fail"])
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v)
            cell.fill = fill; cell.font = N_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    widths = [8,30,14,12,13,14,12,13,8,10,10,13,13,12,12,10,60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save("yfinance_results.xlsx")
    print("\n  📊 Excel saved: yfinance_results.xlsx")


if __name__ == "__main__":
    print(f"\n{'='*60}")
    print(f"  YFINANCE TEST  ·  {len(TICKERS)} tickers  ·  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'='*60}\n")

    results = []
    for i, (name, (ticker, verdict, delisted)) in enumerate(TICKERS.items(), 1):
        print(f"[{i:02d}/{len(TICKERS)}] {ticker}")
        r = test_ticker(name, ticker, verdict, delisted)
        results.append(r)
        print_result(r)
        print()
        if not delisted:
            time.sleep(0.3)

    ok = sum(1 for r in results if r["status"]=="ok")
    partial = sum(1 for r in results if r["status"]=="partial")
    delisted = sum(1 for r in results if r["status"]=="delisted")
    fail = sum(1 for r in results if r["status"]=="fail")

    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"  {GREEN}✅ OK          {ok:>3}{RESET}  — yfinance covers financial metrics")
    print(f"  {YELLOW}⚠️  PARTIAL     {partial:>3}{RESET}  — add missing metrics to pre-curated dataset")
    print(f"  {CYAN}†  DELISTED    {delisted:>3}{RESET}  — extract M&A multiples from SEC EDGAR")
    print(f"  {RED}❌ FAIL        {fail:>3}{RESET}  — check ticker format (see notes below)")
    print(f"  {'─'*40}")
    print(f"     Total        {len(results):>3}")

    if fail:
        print(f"\n  TICKER FORMAT ALTERNATIVES TO TRY:")
        format_hints = {
            "BRETEC.HE":  "→ try 0TN.MU (Munich) or 0TN.F (Frankfurt) as fallback",
            "SECT-B.ST":  "→ try SECTB.ST",
            "AMBU-B.CO":  "→ try AMBUB.CO",
            "BACTI-B.ST": "→ try BACTIB.ST",
            "EKTA-B.ST":  "→ try EKTAB.ST",
            "ALMDT.PA":   "→ try MDT.PA or search Euronext Growth",
            "IBAB.BR":    "→ try IBAB.BT",
            "ELN.MI":     "→ try ELN.MI or ELEN.MI",
        }
        for r in results:
            if r["status"] == "fail":
                hint = format_hints.get(r["ticker"], "→ check exchange suffix")
                print(f"     · {r['ticker']:<15} {hint}")
    print(f"{'='*60}\n")
    save_excel(results)
